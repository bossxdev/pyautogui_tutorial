from openpyxl import load_workbook
from datetime import datetime

from src.constants.constant import DEFAULT_COLUMN

import os

class ExcelOperations:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = None
        self.sheet = None

    def load_workbook(self):
        """โหลด workbook และ active sheet"""
        try:
            self.wb = load_workbook(self.file_path)
            self.sheet = self.wb.active
            return True
        except Exception as e:
            print(f"เกิดข้อผิดพลาดในการโหลดไฟล์: {str(e)}")
            return False

    def update_multiple_cells(self, updates):
        """อัพเดทข้อมูลหลายเซลล์พร้อมกัน"""
        try:
            if not self.wb:
                if not self.load_workbook():
                    return False

            for update in updates:
                self.sheet.cell(
                    row=update["row"], 
                    column={DEFAULT_COLUMN},
                    value=update["value"]
                )
            return True
        except Exception as e:
            print(f"เกิดข้อผิดพลาดในการอัพเดทข้อมูล: {str(e)}")
            return False

    def export_excel(self, export_path=None):
        """Export ไฟล์ Excel ไปยังตำแหน่งที่กำหนด"""
        try:
            if not self.wb:
                if not self.load_workbook():
                    return False

            # สร้างชื่อไฟล์ export ถ้าไม่ได้ระบุ
            if not export_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_name = os.path.splitext(os.path.basename(self.file_path))[0]
                export_path = f"{file_name}_exported_{timestamp}.xlsx"

            # สร้างโฟลเดอร์ exports ถ้ายังไม่มี
            export_dir = "exports"
            if not os.path.exists(export_dir):
                os.makedirs(export_dir)

            # รวม path
            full_export_path = os.path.join(export_dir, export_path)
            
            # บันทึกไฟล์
            self.wb.save(full_export_path)
            print(f"ไฟล์ถูก export ไปยัง: {full_export_path}")
            return True

        except Exception as e:
            print(f"เกิดข้อผิดพลาดในการ export ไฟล์: {str(e)}")
            return False

    def save_and_close(self):
        """บันทึกการเปลี่ยนแปลงและปิดไฟล์"""
        try:
            if self.wb:
                self.wb.save(self.file_path)
                self.wb.close()
                print(f"บันทึกการเปลี่ยนแปลงในไฟล์ {self.file_path} เรียบร้อยแล้ว")
                return True
        except Exception as e:
            print(f"เกิดข้อผิดพลาดในการบันทึกไฟล์: {str(e)}")
            return False

def main():
    file_path = "[NSS-POSNewGen] Script Automate.xlsx"
    excel_ops = ExcelOperations(file_path)

    # ข้อมูลที่ต้องการอัปเดต
    updates = [
        {"row": 667, "value": "Updated test v.4"},
        {"row": 740, "value": "Updated v.5"},
    ]

    # อัพเดทข้อมูล
    if excel_ops.update_multiple_cells(updates):
        # Export ไฟล์
        excel_ops.export_excel()
        # บันทึกการเปลี่ยนแปลงในไฟล์ต้นฉบับ
        excel_ops.save_and_close()

if __name__ == "__main__":
    main()