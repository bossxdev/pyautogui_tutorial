from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime
import os

from src.constants.constant import PATH, NSS_EXCEL_FILE, NSS_WORKSHEET_NAME, NSS_COMMENT_COLUMN


def update_cells(file_path, result):
    """
    อัพเดทข้อมูลเซลล์ในไฟล์ Excel
    """
    wb = load_workbook(file_path)
    ws = wb[NSS_WORKSHEET_NAME]

    if not ws:
        return False

    column = column_index_from_string("N")

    try:
        for results in result:
            for row in ws.iter_rows(column):
                for cell in row:
                    if cell.value == results["cell"]:
                        ws.cell(row=cell.row, column=NSS_COMMENT_COLUMN).value = results["value"]
                        return wb

        return None
    except Exception as e:
        print(f"เกิดข้อผิดพลาดในการอัพเดทข้อมูล: {str(e)}")
        return None


def save_export(workbook, file_path):
    """
    Export ไฟล์ Excel พร้อมระบุชื่อไฟล์
    """
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = os.path.splitext(os.path.basename(file_path))[0]
        export_path = f"{file_name}_exported_{timestamp}.xlsx"

        export_dir = f"{PATH['EXPORT']}"
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)

        full_export_path = os.path.join(export_dir, export_path)

        workbook.save(full_export_path)
        workbook.close()
        print(f"บันทึกการเปลี่ยนแปลงและปิดไฟล์เรียบร้อยแล้ว")

        return True
    except Exception as e:
        print(f"เกิดข้อผิดพลาดในการบันทึกไฟล์: {str(e)}")
        return False


def excel_editor(result):
    """
    ฟังก์ชันสำหรับแก้ไขข้อมูลในไฟล์ Excel และทำการ export
    :param file_path: เส้นทางของไฟล์ Excel
    :param updates: รายการอัพเดทข้อมูลในเซลล์
    """
    file_path = f"{PATH['ASSET']}{NSS_EXCEL_FILE}"
    workbook = update_cells(file_path, [result])

    if workbook:
        save_export(workbook, file_path)
